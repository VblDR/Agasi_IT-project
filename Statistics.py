from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
import Parameters as Parameters
from openpyxl import Workbook, load_workbook
import pyqtgraph as pg
import Exc_Work as Exc


class FormWidget3(QWidget):

    def __init__(self, parent):

        super(FormWidget3, self).__init__(parent)
        self.setFont(QFont('Century Gothic', 15))
        self.setWindowIcon(QIcon('logo.png'))

        self.windowW = Parameters.ParameterSize().ww()

        self.art_line = QLineEdit()
        self.art_line.setPlaceholderText('Введите артикул')

        # массив с датами, имеющимися на данный момент в отчёте
        self.date = ['январь 2016', 'февраль 2016', 'март 2016', 'апрель 2016', 'май 2016', 'июнь 2016', 'июль 2016',
                     'август 2016', 'сентябрь 2016', 'октябрь 2016', 'ноябрь 2016', 'декабрь 2016', 'январь 2017',
                     'февраль 2017', 'март 2017', 'апрель 2017', 'май 2017', 'июнь 2017', 'июль 2017',
                     'август 2017', 'сентябрь 2017', 'октябрь 2017', 'ноябрь 2017', 'декабрь 2017', 'январь 2018',
                     'февраль 2018', 'март 2018', 'апрель 2018', 'май 2018', 'июнь 2018', 'июль 2018',
                     'август 2018', 'сентябрь 2018', 'октябрь 2018', 'ноябрь 2018', 'декабрь 2018']

        # окна выбора периода, за который формируется отчёт
        self.date_line1 = QComboBox(self)
        self.date_line1.addItems(self.date)
        self.date_line1.activated[str].connect(self.on_activated)
        self.date_line2 = QComboBox(self)

        # группируем компоненты в окне
        hbox = QHBoxLayout()
        hbox.addWidget(self.date_line1)
        hbox.addWidget(self.date_line2)

        # поле ввода артикула
        self.artic_line = QLineEdit()
        self.artic_line.setPlaceholderText('Введите артикул')

        # заполняем автодополнение артикулов
        self.artic_line_comp = []
        with open('all_art.txt', 'r') as f:
            self.artic_line_comp = [p.strip() for p in f]
        self.completer1 = QCompleter(self.artic_line_comp, self.artic_line)
        self.artic_line.setCompleter(self.completer1)

        # кнопка вывода статистики
        stat_but = QPushButton('Статистика')
        stat_but.clicked.connect(self.statistic)

        hbox1 = QHBoxLayout()
        hbox1.addWidget(self.artic_line)
        hbox1.addWidget(stat_but)

        # таблица со статистикой
        self.tab1 = QTableWidget()
        self.tab1.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # кнопки для создания графика и экспорта статистики в документ-Excel
        plot_but = QPushButton('График')
        plot_but.clicked.connect(self.plots)
        exp_but = QPushButton('Экспортировать')
        exp_but.clicked.connect(self.createDoc1)

        # группировка компонентов в окне
        hbox2 = QHBoxLayout()
        hbox2.addWidget(plot_but)
        hbox2.addWidget(exp_but)

        vbox = QVBoxLayout()
        vbox.addLayout(hbox)
        vbox.addLayout(hbox1)
        vbox.addWidget(self.tab1)
        vbox.addLayout(hbox2)

        self.setLayout(vbox)

    # меняем выбор даты
    def on_activated(self, text):
        self.date_line2.clear()
        self.date_line2.addItems(self.date[self.date.index(text):])

    # ищем информацию в базе данных
    def statistic(self):
        # проверка правильности введеных данных
        if self.date_line2.currentText() is '':

            warning = QMessageBox()
            warning.setIcon(QMessageBox.Critical)
            warning.setText("Возникла ошибка поиска. Вы не выбрали промежуток времени.")
            warning.setWindowTitle("Ошибка")
            warning.setStandardButtons(QMessageBox.Cancel)

            rslt = warning.exec()

        elif self.artic_line.text() is '':

            warning = QMessageBox()
            warning.setIcon(QMessageBox.Critical)
            warning.setText("Возникла ошибка поиска. Вы не ввели артикул.")
            warning.setWindowTitle("Ошибка")
            warning.setStandardButtons(QMessageBox.Cancel)

            rslt = warning.exec()

        else:
            exbook = load_workbook('stat.xlsx', data_only=True)
            page = exbook.active
            art = self.artic_line.text()
            m1 = self.date.index(self.date_line1.currentText())
            m2 = self.date.index(self.date_line2.currentText())

            indx = []

            for i in range(2, 3044):
                articul2 = page.cell(row=i, column=1).value
                if art == articul2:

                    indx.append(i)

            sale_list = []
            arts = []

            columns = [3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31, 33, 35, 37, 39, 41, 43, 45, 47, 49, 51,
                       53, 55, 57, 59, 61, 63, 65, 67, 69, 71, 73]

            for i in indx:
                arts.append(page.cell(row=i, column=1).value)
                for j in range(m1, m2+1):
                    sale_list.append(page.cell(row=i, column=columns[j]).value)

            self.tab1.setColumnCount(1+m2-m1+1)

            for i in range(len(arts)):
                self.tab1.setRowCount(i + 1)
                item = QTableWidgetItem(arts[i])
                item.setTextAlignment(Qt.AlignLeft)
                self.tab1.setItem(i, 0, item)

            for i in range(len(sale_list)):
                    if i == 0:
                        self.tab1.setColumnWidth(i, self.windowW * 0.2)
                        self.tab1.setHorizontalHeaderLabels(["Артикул"] + [self.date[j] for j in range(m1, m2+1)])
                    item = QTableWidgetItem(str(sale_list[i]))
                    item.setTextAlignment(Qt.AlignHCenter)
                    self.tab1.setItem(0, i+1, item)

            self.art_plot = arts
            self.sale_plot = sale_list
            self.dates = [self.date[j] for j in range(m1, m2+1)]

    def plots(self):
        if self.tab1.item(1, 1) is not '':

            self.plot1 = Plots()
            self.plot1.get_data( self.sale_plot, self.dates)
            self.plot1.show()

        else:
            warning = QMessageBox()
            warning.setIcon(QMessageBox.Critical)
            warning.setText("Вы не выгрузили статистику в таблицу.")
            warning.setWindowTitle("Ошибка")
            warning.setStandardButtons(QMessageBox.Cancel)

            rslt = warning.exec()

    def createDoc1(self):

        if self.tab1.item(0, 0) is not None:

            # окна выбора пути и имени расположения файла
            way = QFileDialog.getSaveFileName(parent=None, caption="Сохранить конфигурацию товаров",
                                              filter="Документ Excel (*.xlsx)")
            filename = way[0]

            m1 = self.date.index(self.date_line1.currentText())
            m2 = self.date.index(self.date_line2.currentText())

        # в случае ошибки
        else:
            warning = QMessageBox()
            warning.setIcon(QMessageBox.Critical)
            warning.setText("Возникла ошибка создания. \nНеобходимо выполнить поиск данных.")
            warning.setWindowTitle("Ошибка")
            warning.setStandardButtons(QMessageBox.Cancel)

            rslt = warning.exec()

            return rslt

        # заполняем созданный Excel-документ
        row = [["Артикул"] + [self.date[j] for j in range(m1, m2+1)]]
        lst = []
        lst.append(self.art_plot[0])
        for i in range(len(self.sale_plot)):
            lst.append(self.sale_plot[i])
        row.append(lst)
        Exc.create(row, filename, 'Статистика')


# окно графика
class Plots(QWidget):

    def __init__(self):
        super(QWidget, self).__init__()

        # настройки окна
        self.setFont(QFont('Century Gothic', 10))
        self.setFont(QFont('Century Gothic', 10))
        self.setWindowTitle('График')
        window_w1 = Parameters.ParameterSize().ww()
        window_h1 = Parameters.ParameterSize().wh()
        layout = QVBoxLayout(self)
        self.setFixedSize(window_w1 * 0.5, window_h1 * 0.5)
        # сам график
        self.view = pg.PlotWidget()

        layout.addWidget(self.view)
        self.setLayout(layout)

    def get_data(self, sale_list, dates):
        # наполняем график данными из статистики
        ticks = [list(zip(range(len(dates)), dates))]
        xdict = dict(enumerate(dates))
        xax = self.view.getAxis('bottom')
        xax.setTicks(ticks)
        self.curve = self.view.plot(list(xdict.keys()), sale_list)
