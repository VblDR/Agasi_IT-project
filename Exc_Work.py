from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QMessageBox


def pusk_but_click(articul, number, check):

    # проверка на правильность введенных данных
    if number != '':

        # если выбраны изделия
        if check == 1:
            lens = articul.find('x')-1
            try:
                numlist, namlist, artlist = what_name2(articul, int(number), lens)
            except:
                numlist, namlist, artlist = 0, 0, 0

            return numlist, namlist, artlist

        # если выбраны эндопротезы
        elif check == 0:
            numlist, namlist, artlist = what_name(articul, int(number))
            return numlist, namlist, artlist

    else:
        warning2 = QMessageBox()
        warning2.setIcon(QMessageBox.Critical)
        warning2.setText("Возникла ошибка. "
                         "\nВы не заполнили поля 'Атикул' или 'Количество' или заполнили их неправильно.")
        warning2.setWindowTitle("Ошибка")
        warning2.setStandardButtons(QMessageBox.Cancel)

        rs = warning2.exec()
        return rs


# эндопротезы
def what_name(art, number):
    # открываем документ с данными по продукции
    exbook = load_workbook('exwork.xlsx', data_only=True)
    page = exbook.active
    indlist = []

    # находим нужные нам артикулы эндопротезов
    if art == '10.016.xx':
        for i in range(2, 1973):
            articul2 = page.cell(row=i, column=1).value
            if articul2[:6] == art[:6]:

                indlist.append(i)

    elif art == 'OM 001.02.xx':
        for i in range(2, 1973):
            articul2 = page.cell(row=i, column=1).value
            if articul2[:9] == art[:9]:

                indlist.append(i)

    namlist = []
    percentlist = []
    artlist = []
    # заполняем массивы артикулами, названиями изделий и процентом продаж
    for i in indlist:
        namlist.append(page.cell(row=i, column=2).value)
        percentlist.append(round(page.cell(row=i, column=10).value, 2))
        artlist.append(page.cell(row=i, column=1).value)
    numlist = []

    # перерасчет процентов с учетом введенного пользователем количества продукции, необходимого для закупки
    for i in percentlist:
        numlist.append(round(i * number))
    if sum(numlist) != number:
        number2 = number - sum(numlist)
        for i in range(number2):
            ind = numlist.index(min(numlist))
            numlist[ind] += 1

    if sum(numlist) >= number:
        number2 = sum(numlist) - number
        for i in range(number2):
            numlist[percentlist.index(min(percentlist))] -= 1
            percentlist[percentlist.index(min(percentlist))] = 9999

    return numlist, namlist, artlist


# изделия
def what_name2(art, number, lens):
    # открываем документ с данными по продукции
    exbook = load_workbook('exwork.xlsx', data_only=True)
    page = exbook.active
    indlist = []

    # находим нужные нам артикулы изделий
    for i in range(2, 1973):
        articul2 = page.cell(row=i, column=1).value
        if articul2[:lens] == art[:lens]:

            indlist.append(i)

    namlist = []
    percentlist = []
    artlist = []

    # заполняем массивы артикулами, названиями изделий и процентом продаж
    for i in indlist:
        namlist.append(page.cell(row=i, column=2).value)
        percentlist.append(round(page.cell(row=i, column=10).value, 2))
        artlist.append(page.cell(row=i, column=1).value)

    numlist = []

    # перерасчет процентов с учетом введенного пользователем количества продукции, необходимого для закупки
    for i in percentlist:
        numlist.append(round(i*number))

    if sum(numlist) != number:
        number2 = number - sum(numlist)
        for i in range(number2):
            ind = numlist.index(min(numlist))
            numlist[ind] += 1

    if sum(numlist) >= number:
        number2 = sum(numlist) - number
        for i in range(number2):
            numlist[percentlist.index(min(percentlist))] -= 1
            percentlist[percentlist.index(min(percentlist))] = 9999

    return numlist, namlist, artlist


def create(row, filename, title):
    # создаем новый Excel-документ
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # заполняем документ полученными данными
    for i in row:
        ws.append(i)

    # сохраняем документ
    wb.save(filename)

    # окно, оповещающее об успешном сохранении
    welldone = QMessageBox()
    welldone.setIcon(QMessageBox.Information)
    welldone.setText("Сохранение прошло успешно.")
    welldone.setWindowTitle("Выполнено!")
    welldone.setStandardButtons(QMessageBox.Ok)

    rslt = welldone.exec()

    return rslt
